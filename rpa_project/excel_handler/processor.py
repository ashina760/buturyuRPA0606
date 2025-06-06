import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from excel_handler.utils import format_date

class ExcelProcessor:
    def __init__(self, file_path):
        """
        初始化处理器

        参数:
            file_path (str): Excel 文件路径
        """
        self.file_path = file_path
        self.workbook = load_workbook(file_path, data_only=True)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        self.workbook_name = os.path.basename(file_path)
        self.sheet_name = self.sheet.title
        self.min_row = None
        self.max_row = None

    def has_multiple_sheets(self):
        """
        判断是否存在多个 sheet

        返回:
            str: 错误信息（如果有多个 sheet）
        """
        if len(self.workbook.sheetnames) > 1:
            return f"err: {self.workbook_name} 有多个表"

    def is_cell_empty(self, cell_refs):
        """
        检查给定的单元格是否为空

        参数:
            cell_refs (list[str]): 要检查的单元格引用，如 ["J6", "A1"]

        返回:
            list[str] | None: 返回所有为空的单元格信息列表，如果都非空则返回 None
        """
        errors = []
        for ref in cell_refs:
            cell = self.sheet[ref]
            if isinstance(cell, tuple):  # 防止用户传了 "A1:A5" 之类的区域
                for c in cell:
                    if c.value is None or str(c.value).strip() == "":
                        errors.append(f"❌ {c.coordinate} 为空")
            else:
                if cell.value is None or str(cell.value).strip() == "":
                    errors.append(f"{ref} 为空")
        return errors if errors else None

    def is_title_valid(self, columns=None, expected_values=None):
        """
        检查指定列范围内的表头是否匹配预期字段。

        参数:
            columns (list[str], optional): 要检查的列字母列表，例如 ["A", "B", "C"]
            expected_values (list[str], optional): 对应每列的预期表头值，例如 ["№", "倉庫", "ＪＡＮ"]

        返回:
            str | None: 如果有列不匹配，返回错误信息，否则返回 None。
        """
        if columns is None:
            columns = list("CDEFGHIJk")  # 默认 A~J
        if expected_values is None:
            expected_values = ["仕入先コード", "入荷倉庫コード", "商品コード", "商品名（伝票用）", "発注数量", "納期", "発注単価", "発注金額"]

        for col, expected_value in zip(columns, expected_values):
            col_idx = column_index_from_string(col)
            if not any(
                self.sheet.cell(row=row, column=col_idx).value == expected_value
                for row in range(1, self.sheet.max_row + 1)
            ):
                return f" {col} 列不是规定的标题 {expected_value}"

        return None

    def get_min_max_row(self, target_column='G', keyword='発注数量'):
        """
        获取数据的起始行与结束行

        参数:
            target_column (str): 搜索关键词的列（例如 "G"）
            keyword (str): 搜索关键词（默认 "発注数量"）

        返回:
            tuple[int, int]: 数据区的最小行号和最大行号
        """
        col_idx = column_index_from_string(target_column)
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=col_idx).value == keyword:
                self.min_row = row + 1
                break
        self.max_row = max(
            (row for row in range(1, self.sheet.max_row + 1)
            if self.sheet.cell(row=row, column=col_idx).value is not None),
            default=0
        )
        return self.min_row, self.max_row

    def delete_empty_rows(self, column):
        """
        删除指定列中为空的所有行

        参数:
            column (str): 要检查空值的列（例如 "H"）

        返回:
            tuple[int, list[int]]: 删除后工作表的最大行号，和被删除的行号列表
        """
        if self.min_row is None or self.max_row is None:
            self.get_min_max_row(column)
        
        col_idx = column_index_from_string(column)
        
        delete_rows = [row for row in range(self.min_row, self.max_row + 1)
                    if self.sheet.cell(row=row, column=col_idx).value is None]
        
        for row in reversed(delete_rows):
            self.sheet.delete_rows(row)
        
        self.max_row = self.sheet.max_row
        return self.sheet.max_row, delete_rows

    def find_empty_cells(self,min_col=3, max_col=11):
        """
        找出数据区域中 min_col~ max_col 列所有空单元格

        返回:
            list[str]: 所有空单元格的位置（例如 ["B5", "J9"]）
        """
        self.get_min_max_row()
        empty_cells = []
        for row in self.sheet.iter_rows(min_row=self.min_row, max_row=self.max_row, min_col= min_col, max_col= max_col):  # B~F
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    empty_cells.append(cell.coordinate)

        return empty_cells

    def save_cleaned_sheet(self, output_path):
        """
        保存清洗后的数据区域到新的 Excel 文件

        参数:
            output_path (str): 新文件保存路径

        返回:
            str: 输出路径
        """
        if self.min_row is None or self.max_row is None:
            self.get_min_max_row()
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = self.sheet_name
        for row in self.sheet.iter_rows(min_row=self.min_row - 1, max_row=self.max_row, values_only=True):
            new_sheet.append(row)
        new_wb.save(output_path)
        return output_path

    def close(self):
        """
        关闭工作簿
        """
        self.workbook.close()

    def save(self, save_path: str = None):
        """
        保存当前工作簿到指定路径。如果未提供路径，则覆盖原文件。

        参数:
            save_path (str, optional): 要保存的路径。默认为初始化时的路径。
        """
        if save_path is None:
            original_filename = os.path.basename(self.file_path)
            directory = os.path.dirname(self.file_path)
            save_path = os.path.join(directory, f"NEW_{original_filename}")

        self.workbook.save(save_path)
        

    def get_column_dates_with_colD(self, date_column='H', id_column='D'):
        """
        获取指定列中的所有有效日期，并将其与 D 列的值配对为元组 (D值, 日期)

        参数:
            date_column (str): 日期所在列，例如 "H"
            id_column (str): 需要配对的 ID 列，例如 "D"

        返回:
            list[tuple[str, str]]: 格式为 (D列的值, 格式化后的日期字符串)
        """
        if self.min_row is None or self.max_row is None:
            self.get_min_max_row(date_column)

        date_col_idx = column_index_from_string(date_column)
        id_col_idx = column_index_from_string(id_column)

        results = []

        for row in range(self.min_row, self.max_row + 1):
            id_val = self.sheet.cell(row=row, column=id_col_idx).value
            date_val = self.sheet.cell(row=row, column=date_col_idx).value

            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y%m%d')
            elif isinstance(date_val, str):
                for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
                    try:
                        date_obj = datetime.strptime(date_val, fmt)
                        date_str = date_obj.strftime('%Y%m%d')
                        break
                    except ValueError:
                        continue
                else:
                    continue
            else:
                continue

            results.append((str(id_val), date_str))

        return results
    
    def create_upload_data(self, save_dir,fill_values):
        """
        生成用于上传的 Excel 数据，只处理当前工作表。

        参数:
            save_dir (str): 输出文件夹路径

        返回:
            str: 保存后的文件路径
        """

        headers = [
            "T", "仕入先コード", "センターコード", "指定納期", "担当者コード", "決裁区分", "決裁番号", "発注残管理",
            "商品コード", "発注数量", "明細備考1", "明細備考2", "決裁営業", "お客様", "伝票備考"
        ]

        os.makedirs(save_dir, exist_ok=True)

        wb_new = Workbook()
        sheet_new = wb_new.active
        sheet_new.title = self.sheet_name
        sheet_new.append(headers)

        for row in self.sheet.iter_rows(min_row=self.min_row, values_only=True):
            new_row = [
                None, row[2], row[3], row[7], None, None, None, None,
                row[4], row[6], None, None, None, None, row[10]
            ]
            sheet_new.append(new_row)

        for row in sheet_new.iter_rows(min_row=2, min_col=4, max_col=4):  # 指定納期列格式化
            for cell in row:
                format_date(cell)

        for col in [2, 3, 9, 10, 15]:  # 一些列转为字符串
            for row in sheet_new.iter_rows(min_row=2, min_col=col, max_col=col):
                for cell in row:
                    if cell.value is not None:
                        cell.value = str(cell.value)

        for row_idx in range(2, sheet_new.max_row + 1):  # 注意从第2行开始
            for col, val in fill_values.items():
                sheet_new.cell(row=row_idx, column=col, value=val)

        # 删除“発注数量”为空或为0的行（第10列）
        rows_to_delete = [
            row[0].row for row in sheet_new.iter_rows(min_row=2, min_col=10, max_col=10)
            if row[0].value in [None, 0, "0"]
        ]
        for row_idx in reversed(rows_to_delete):
            sheet_new.delete_rows(row_idx)

        save_path = os.path.join(save_dir, "nagashikomi.xlsx")
        wb_new.save(save_path)
        return save_path
    
    def get_column_based_dict(self):
        """
        将表格按列转换为字典格式：
        - 第一行为字段名
        - 第二行起为对应字段的值列表

        返回:
            dict: {字段名: [值1, 值2, ...], ...}
        """
        rows = list(self.sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return {}

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows = rows[1:]

        column_dict = {header: [] for header in headers}

        for row in data_rows:
            for idx, value in enumerate(row):
                if idx < len(headers):
                    column_dict[headers[idx]].append(value)

        return column_dict
    
    def get_cell_values_from_workbook(self, cell_list):
        """
        从 workbook 中指定 sheet 获取指定单元格的值，并以字符串形式返回

        :param cell_list: 单元格地址列表（如 ["C2", "J6"]）
        :return: 包含对应单元格文字列的列表
        """
        sheet = self.sheet
        return [str(sheet[cell].value).strip() if sheet[cell].value is not None else "" for cell in cell_list]

    def convert_column_to_yyyymmdd(self, column_letter, start_row=2):
        """
        将指定列的所有单元格内容转为 'yyyymmdd' 字符串格式

        参数:
            column_letter (str): 要转换的列，例如 'C'
            start_row (int): 从哪一行开始处理（默认跳过表头）
        """
        col_idx = column_index_from_string(column_letter)
        for row in range(start_row, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=col_idx)
            value = cell.value
            if value is None:
                continue
            try:
                # 如果是 datetime 类型，直接格式化
                if isinstance(value, datetime):
                    cell.value = value.strftime('%Y%m%d')
                else:
                    # 尝试将字符串转换为日期
                    parsed = datetime.strptime(str(value), '%Y/%m/%d')
                    cell.value = parsed.strftime('%Y%m%d')
            except Exception as e:
                print(f"⚠️ 第 {row} 行转换失败，原值: {value}, 错误: {e}") 

























    

    # def check_delivery_date_match(self, wb2_path, col1="D", col2="H"):
    #     """
    #     检查交货日期是否匹配参考表格，并验证是否为过去日期

    #     参数:
    #         wb2_path (str): 第二个工作簿的文件路径
    #         col1 (str): 用于匹配列名的列（当前 sheet）
    #         col2 (str): 要验证是否匹配的数据列（当前 sheet）

    #     返回:
    #         str | None: 匹配错误信息或 None
    #     """
    #     if self.min_row is None:
    #         self.get_min_max_row(col2)

    #     wb2 = openpyxl.load_workbook(wb2_path, data_only=True)
    #     sheet2 = wb2.active
    #     wb2_header = [cell.value for cell in sheet2[1] if cell.value is not None]

    #     col1_index = column_index_from_string(col1) - 1
    #     col2_index = column_index_from_string(col2) - 1

    #     found = False
    #     err_msg = None

    #     for row in self.sheet.iter_rows(min_row=self.min_row, values_only=True):
    #         val1 = row[col1_index]
    #         val2 = row[col2_index]
    #         val2_parsed = datetime.strptime(val2, "%Y/%m/%d")
    #         if val1 in wb2_header:
    #             match_col_index = wb2_header.index(val1) + 1
    #             col_values = [
    #                 sheet2.cell(row=r, column=match_col_index).value
    #                 for r in range(2, sheet2.max_row + 1)
    #             ]
    #             col_values = [v for v in col_values if v is not None]

    #             if val2_parsed not in col_values:
    #                 print(f"❌ 未找到匹配项: {col2}列值 {val2} 不在 wb2 中 {val1} 对应列中")
    #                 found = True
    #                 break

    #         if isinstance(val2, str):
    #             try:
    #                 val2_parsed = datetime.strptime(val2, "%Y/%m/%d")
    #                 today = datetime.today()
    #                 if val2_parsed <= today:
    #                     print(f"⚠️ 发现 {col2}列值 {val2_parsed.strftime('%Y/%m/%d')} 是今天或更早的日期")
    #                     found = True
    #                     break
    #             except ValueError:
    #                 pass

    #     wb2.close()

    #     if found:
    #         err_msg = f"!!Data is Mistaken - {val2}"
    #     return err_msg
    
