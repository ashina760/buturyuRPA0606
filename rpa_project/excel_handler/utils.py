from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from openpyxl.utils import column_index_from_string
from datetime import datetime
import os

def check_dates_in_dict(id_date_tuple, delivery_date_dic):
    unmatched = []

    for _id, date_str in id_date_tuple:
        date_str = date_str.strip()  # 去除空格
        delivery_dates = delivery_date_dic.get(_id, [])

        valid_dates = [
            d.strip() for d in delivery_dates
            if isinstance(d, str) and d.strip() != '0'
        ]

        if date_str not in valid_dates:
            unmatched.append((_id, date_str))

    return unmatched

def check_past_dates(id_date_tuple):
    """
    检查 id_date_tuple 中每个元组的日期是否早于今天

    参数:
        id_date_tuple (list[tuple[str, str]]): (id, date_str) 组成的列表，date_str 格式为 yyyymmdd

    返回:
        list[tuple[str, str]]: 所有日期早于今天的 (id, date_str) 元组
    """
    today = int(datetime.today().strftime("%Y%m%d"))
    past_dates = []

    for _id, date_str in id_date_tuple:
        try:
            date_int = int(date_str)
            if date_int < today:
                past_dates.append((_id, date_str))
        except ValueError:
            continue  # 忽略非法日期字符串

    return past_dates

def format_date(cell):
    """格式化日期为 yyyy/mm/dd 格式"""
    if isinstance(cell.value, datetime):
        cell.value = cell.value.strftime('%Y/%m/%d')
    elif isinstance(cell.value, str):

        date_obj = datetime.strptime(cell.value, '%Y-%m-%d')
        cell.value = date_obj.strftime('%Y/%m/%d')

def format_column_to_yyyymmdd(sheet: Worksheet, column: str, start_row: int = 2):
    """
    将指定列的所有日期值统一格式化为 'yyyymmdd' 字符串。

    参数:
        sheet (Worksheet): 要处理的工作表
        column (str): 列标，如 'H'
        start_row (int): 起始行号，默认为 2（跳过标题）

    返回:
        list[int]: 被格式化的行号列表
    """
    formatted_rows = []
    col_idx = column_index_from_string(column)

    for row in range(start_row, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col_idx)
        if cell.value is None:
            continue
        try:
            # 如果是 datetime 类型
            if isinstance(cell.value, datetime):
                cell.value = cell.value.strftime('%Y%m%d')
                formatted_rows.append(row)
            # 如果是字符串，尝试转为 datetime 对象
            elif isinstance(cell.value, str):
                # 兼容各种常见格式
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
                    try:
                        date_obj = datetime.strptime(cell.value, fmt)
                        cell.value = date_obj.strftime('%Y%m%d')
                        formatted_rows.append(row)
                        break
                    except ValueError:
                        continue
        except Exception as e:
            print(f"❌ 第 {row} 行格式化失败: {cell.value} - {e}")

    return formatted_rows

def get_latest_file(download_dir: str) -> str:
    """
    获取指定文件夹中最新（最近修改）的文件路径。

    参数:
        download_dir (str): 下载文件夹路径

    返回:
        str: 最新文件的完整路径（如果没有文件则返回 None）
    """
    files = [os.path.join(download_dir, f) for f in os.listdir(download_dir)]
    files = [f for f in files if os.path.isfile(f)]

    if not files:
        return None

    latest_file = max(files, key=os.path.getmtime)
    return latest_file

if __name__ == "__main__":
    from openpyxl import load_workbook
    wb = load_workbook("newinput.xlsx")
    sheet = wb.active

    formatted_rows = format_column_to_yyyymmdd(sheet, "H")
    print("已格式化行：", formatted_rows)

    wb.close


